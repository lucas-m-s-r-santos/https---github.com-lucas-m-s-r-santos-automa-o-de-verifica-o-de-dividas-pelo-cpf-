import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
planilha = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha['Sheet1']

for l in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = l
    
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    sleep(2)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(2)
    botao_pesquisa = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(2)
    botao_pesquisa.click()
    sleep(4)
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    if status.text == 'em dia':
        
        data_pag = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
        
        data_pag_limpo =data_pag.text.split()[3]
        metodo_limpo = metodo.text.split()[3]
        
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pag_limpo, metodo_limpo])
        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')